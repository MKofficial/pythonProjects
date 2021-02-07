#!/usr/bin/env python

import csv
import os

import openpyxl as opx
from openpyxl.styles import PatternFill, Side, Border

__author__ = "Matthew Kriz"
__copyright__ = "Copyright 2020, DDucks nTrophy project"
__credits__ = ["Matthew Kriz"]
__license__ = "GNU General Public License v3.0"
__version__ = "1.0.1"
__maintainer__ = "Matthew Kriz"
__email__ = "matej.kriz2003@gmail.com"
__status__ = "Published"


class Table:
    """
    Creating, adjusting, coloring excel tables.
    Creating maps from zero/one xlsx files.
    """
    workbook = None
    sheet = None
    workbook_name = None

    def __init__(self, workbook_name: str):
        if not os.path.exists(workbook_name):
            self.workbook = opx.Workbook()
        else:
            self.workbook = opx.load_workbook(workbook_name)

        self.sheet = self.workbook.active
        self.workbook_name = workbook_name

        self.workbook.save(workbook_name)

    @staticmethod
    def convertCsvToExcel(csv_file: str) -> None:
        """
        Not quite functional method yet! Excel has some problem to display the number fetched from csv file.
        :param csv_file: File or Path that leads to csv file.
        :return: None
        """
        workbook = opx.Workbook()
        sheet = workbook.active

        with open(csv_file, 'r') as file:
            for row in csv.reader(file):
                sheet.append(row)

        workbook.save(csv_file.removesuffix('.csv') + '.xlsx')

    @staticmethod
    def colorTable(workbook_name: str, color: str = 'FF0000') -> None:
        """
        Color every cell having 1 as a value.
        :param workbook_name: Name of workbook which will be colored
        :param color: Specific color which will be used
        :return: None
        """
        workbook = opx.open(workbook_name)
        sheet = workbook.active

        for row in range(1, sheet.max_row + 1):
            for cell in sheet[row]:
                if cell.value == 1:
                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')

        workbook.save(workbook_name.removesuffix('.xlsx') + '_adjust.xlsx')

    @staticmethod
    def adjustCells(workbook_name: str, height: int = 50, width: int = 50) -> None:
        """
        Adjust all cells to specific width and height.
        :param workbook_name: Name of workbook which will be adjusted
        :param height: Height of cell
        :param width: Width of cell
        :return: None
        """
        workbook = opx.open(workbook_name)
        sheet = workbook.active

        for i in range(1, sheet.max_row):
            sheet.row_dimensions[i].height = height

        for j in range(1, sheet.max_column):
            sheet.column_dimensions.width = width

        workbook.save(workbook_name)

    def fillZeros(self) -> None:
        """
        Fill all cells with zeros. Use for resetting values.
        :return: None
        """
        for row in range(1, self.sheet.max_row + 1):
            for column in range(1, self.sheet.max_column + 1):
                self.sheet.cell(row, column, 0)

        self.workbook.save(self.workbook_name)

    def resetColorAndBorders(self) -> None:
        """
        Reset all cell colors and all borders to its initial state.
        :return: None
        """
        color = '00ffffff'
        thin = Side(border_style='thin', color='f2f2f2')

        for row in range(1, self.sheet.max_row + 1):
            for column in range(1, self.sheet.max_column + 1):
                self.sheet.cell(row, column).fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                self.sheet.cell(row, column).border = Border(top=thin, right=thin, bottom=thin, left=thin)

        self.workbook.save(self.workbook_name)

    def resetColor(self):
        """
        Reset all cells colors to its initial state.
        :return:
        """
        color = '00ffffff'

        for row in range(1, self.sheet.max_row + 1):
            for column in range(1, self.sheet.max_column + 1):
                self.sheet.cell(row, column).fill = PatternFill(start_color=color, end_color=color, fill_type='solid')

        self.workbook.save(self.workbook_name)

    def resetBorders(self):
        """
        Reset all cells borders to its initial state.
        :return:
        """
        thin = Side(border_style='thin', color='f2f2f2')

        for row in range(1, self.sheet.max_row + 1):
            for column in range(1, self.sheet.max_column + 1):
                self.sheet.cell(row, column).border = Border(top=thin, right=thin, bottom=thin, left=thin)

        self.workbook.save(self.workbook_name)

    @staticmethod
    def copyValues(workbook_from: str, workbook_to: str) -> None:
        """
        Copy values from one workbook to another. If a value of a cell in copied workbook is equal to 1 then a cell with
        the same position in the second workbook will copy its value and color.
        :param workbook_from: First workbook which will 'share' values.
        :param workbook_to: Second workbook which will 'obtain' values.
        :return: None
        """
        wb_from = opx.load_workbook(workbook_from)
        sheet_from = wb_from.active

        wb_to = opx.load_workbook(workbook_to)
        sheet_to = wb_to.active

        for row in range(1, sheet_from.max_row + 1):
            for column in range(1, sheet_from.max_column + 1):

                if sheet_from.cell(row, column).value == 1:
                    cell = sheet_to.cell(row, column, 1)
                    cell_color = cell.fill.start_color.index
                    cell.fill = PatternFill(start_color=cell_color, end_color=cell_color, fill_type='solid')

        wb_to.save(workbook_to)

    def deleteValues(self):
        """
        Delete all values in workbook.
        NO RECOVERY ONCE IS DATA DELETED. USE THIS FUNCTION CAREFULLY.
        :return: None
        """
        for row in range(1, self.sheet.max_row + 1):
            for column in range(1, self.sheet.max_column + 1):
                self.sheet.cell(row, column, '')

        # DO NOT USE THE SAME NAME AS 'woorkbook_name'. NO RECOVERY.
        self.workbook.save('vysledna_mapa.xlsx')
